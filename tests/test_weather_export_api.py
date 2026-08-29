from io import BytesIO

from openpyxl import load_workbook

from app import create_app
from config import Settings
from geocoding import CityResolution
from parser import City
from weather_client import WeatherData


class FakeWeatherClient:
    def __init__(self):
        self.forecast_calls = []

    def get_current(self, city):
        return WeatherData(22.0, "晴", 50, 2.0, False)

    def get_forecast(self, city, day_offset):
        self.forecast_calls.append((city.name, day_offset))
        return WeatherData(20.0 + day_offset, "雨", 70 + day_offset, 3.0, True)


class FakeResolver:
    def resolve(self, location):
        if location == "长沙":
            return CityResolution(City("长沙", 28.2282, 112.9388))
        return None


def local_app():
    return create_app(settings=Settings(), weather_client=FakeWeatherClient())


def test_follow_up_exports_last_weather_result_as_download():
    http = local_app().test_client()
    http.post(
        "/chat",
        json={"message": "深圳明天天气怎么样", "session_id": "export-follow-up"},
    )

    response = http.post(
        "/chat",
        json={"message": "把刚才天气导出为 Excel", "session_id": "export-follow-up"},
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body["intent"] == "weather_export"
    assert body["export"]["format"] == "xlsx"
    download = http.get(body["export"]["download_url"])
    assert download.status_code == 200
    assert download.headers["Cache-Control"] == "no-store"
    assert "attachment" in download.headers["Content-Disposition"]
    workbook = load_workbook(BytesIO(download.data))
    assert workbook.active["A2"].value == "深圳"


def test_query_and_export_can_happen_in_one_message():
    http = local_app().test_client()

    response = http.post(
        "/chat",
        json={"message": "把深圳和广州明天天气导出为 PDF", "session_id": "export-inline"},
    )

    body = response.get_json()
    assert response.status_code == 200
    assert body["intent"] == "weather_export"
    assert body["export"]["format"] == "pdf"
    assert "cities" not in body
    assert "weather" not in body
    assert "results" not in body
    assert http.get(body["export"]["download_url"]).data.startswith(b"%PDF")


def test_ordered_itinerary_uses_a_different_day_for_each_city_and_only_returns_file():
    weather = FakeWeatherClient()
    app = create_app(
        settings=Settings(),
        weather_client=weather,
        city_resolver=FakeResolver(),
    )
    http = app.test_client()

    response = http.post(
        "/chat",
        json={
            "message": (
                "依次帮我查看，北京，深圳，广州，长沙，的天气，结果输出excel表，"
                "我目前居住在杭州，还需要告诉我出行需要带什么，每个城市待一天"
            ),
            "session_id": "ordered-itinerary",
        },
    )

    assert response.status_code == 200
    body = response.get_json()
    assert weather.forecast_calls == [
        ("北京", 1),
        ("深圳", 2),
        ("广州", 3),
        ("长沙", 4),
    ]
    assert body["intent"] == "weather_export"
    assert body["display_mode"] == "text"
    assert set(body) == {"session_id", "intent", "display_mode", "answer", "export"}

    download = http.get(body["export"]["download_url"])
    workbook = load_workbook(BytesIO(download.data))
    report = workbook["天气报告"]
    assert [report.cell(row=row, column=1).value for row in range(2, 6)] == [
        "北京",
        "深圳",
        "广州",
        "长沙",
    ]
    assert len({report.cell(row=row, column=2).value for row in range(2, 6)}) == 4
    assert all(report.cell(row=row, column=8).value for row in range(2, 6))
    assert "出行清单" in workbook.sheetnames


def test_unordered_trip_exports_each_city_for_all_trip_days():
    weather = FakeWeatherClient()
    app = create_app(settings=Settings(), weather_client=weather)
    http = app.test_client()

    response = http.post(
        "/chat",
        json={
            "message": "北京和深圳出差3天，结果输出Excel",
            "session_id": "unordered-itinerary",
        },
    )

    assert response.status_code == 200
    assert weather.forecast_calls == [
        ("北京", 1),
        ("北京", 2),
        ("北京", 3),
        ("深圳", 1),
        ("深圳", 2),
        ("深圳", 3),
    ]
    body = response.get_json()
    workbook = load_workbook(BytesIO(http.get(body["export"]["download_url"]).data))
    assert workbook["天气报告"].max_row == 7


def test_export_without_format_asks_for_format():
    http = local_app().test_client()
    http.post(
        "/chat",
        json={"message": "北京今天天气", "session_id": "export-no-format"},
    )

    response = http.post(
        "/chat",
        json={"message": "把它导出来", "session_id": "export-no-format"},
    )

    assert response.status_code == 200
    assert "Word、Excel、PDF 或 Markdown" in response.get_json()["answer"]


def test_export_without_weather_snapshot_does_not_invent_data():
    response = local_app().test_client().post(
        "/chat",
        json={"message": "导出为 PDF", "session_id": "export-empty"},
    )

    assert response.status_code == 200
    assert "先查询天气" in response.get_json()["answer"]


def test_unknown_export_id_returns_404():
    response = local_app().test_client().get("/api/exports/not-a-real-id")

    assert response.status_code == 404
